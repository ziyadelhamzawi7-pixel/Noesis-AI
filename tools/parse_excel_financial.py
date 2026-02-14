"""
Enhanced Excel parser for financial model analysis.
Extracts formulas, cell dependencies, sheet types, and financial structure.
"""

import sys
import io
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Set
import json
from loguru import logger

try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    logger.error("Required packages not installed. Run: pip install pandas openpyxl")
    sys.exit(1)


class FinancialExcelParser:
    """
    Enhanced Excel parser for financial models.
    Extracts formulas, cell relationships, and financial structure.
    """

    # Maximum number of cells to scan for formula extraction
    # Prevents hanging on large sparse Excel files
    # Lowered from 50k to 20k due to slow iter_rows() in read_only mode
    MAX_CELLS_FOR_FORMULA_SCAN = 20000  # ~20k cells max

    # All supported ISO 4217 currency codes (used in detection patterns)
    ALL_CURRENCY_CODES = [
        'USD', 'CAD', 'BRL', 'MXN', 'ARS', 'CLP', 'COP', 'PEN',
        'EUR', 'GBP', 'CHF', 'SEK', 'NOK', 'DKK', 'PLN', 'CZK', 'HUF', 'RUB', 'TRY', 'RON', 'BGN', 'HRK', 'UAH',
        'AED', 'SAR', 'QAR', 'KWD', 'BHD', 'OMR', 'JOD', 'ILS', 'IQD', 'IRR', 'LBP',
        'JPY', 'CNY', 'INR', 'KRW', 'TWD', 'HKD', 'SGD', 'AUD', 'NZD', 'THB', 'MYR', 'IDR', 'PHP', 'VND', 'PKR', 'BDT', 'LKR', 'MMK', 'KHR',
        'NGN', 'GHS', 'KES', 'TZS', 'UGX', 'ZAR', 'EGP', 'MAD', 'XOF', 'XAF', 'ETB', 'RWF', 'ZMW', 'MWK', 'BWP', 'MZN', 'AOA', 'CDF', 'DZD', 'TND', 'LYD', 'SDG',
    ]

    # Keywords for detecting sheet types
    PL_KEYWORDS = [
        'revenue', 'sales', 'cogs', 'cost of goods', 'gross profit', 'gross margin',
        'operating', 'opex', 'ebitda', 'ebit', 'net income', 'net profit',
        'depreciation', 'amortization', 'interest expense', 'tax'
    ]

    BS_KEYWORDS = [
        'assets', 'liabilities', 'equity', 'cash', 'accounts receivable',
        'accounts payable', 'inventory', 'property', 'plant', 'equipment',
        'debt', 'loans', 'retained earnings', 'shareholders'
    ]

    CF_KEYWORDS = [
        'cash flow', 'operating activities', 'investing activities',
        'financing activities', 'net cash', 'free cash flow', 'fcf',
        'capex', 'capital expenditure', 'working capital'
    ]

    SAAS_KEYWORDS = [
        'arr', 'mrr', 'annual recurring', 'monthly recurring', 'churn',
        'cac', 'ltv', 'lifetime value', 'customer acquisition', 'arpu',
        'net revenue retention', 'nrr', 'gross revenue retention'
    ]

    ASSUMPTIONS_KEYWORDS = [
        'assumptions', 'inputs', 'parameters', 'drivers', 'scenarios',
        'sensitivity', 'variables'
    ]

    # Pre-compiled currency detection patterns (compiled once at class definition time)
    _CURRENCY_PATTERNS = {
        # Americas
        'USD': re.compile(r'US\s*\$|\bUSD\b|\bUS\s+dollars?', re.IGNORECASE),
        'CAD': re.compile(r'\bCAD\b|C\$|canadian\s*dollars?', re.IGNORECASE),
        'BRL': re.compile(r'R\$|\bBRL\b|\breais\b|\breals?\b', re.IGNORECASE),
        'MXN': re.compile(r'\bMXN\b|mexican\s*pesos?', re.IGNORECASE),
        'ARS': re.compile(r'\bARS\b|argentine\s*pesos?', re.IGNORECASE),
        'CLP': re.compile(r'\bCLP\b|chilean\s*pesos?', re.IGNORECASE),
        'COP': re.compile(r'\bCOP\b|colombian\s*pesos?', re.IGNORECASE),
        'PEN': re.compile(r'\bPEN\b|\bsoles?\b|nuevos?\s*soles?', re.IGNORECASE),
        # Europe
        'EUR': re.compile(r'\u20ac|\bEUR\b|\beuros?\b', re.IGNORECASE),
        'GBP': re.compile(r'\u00a3|\bGBP\b|pounds?\s*sterling', re.IGNORECASE),
        'CHF': re.compile(r'\bCHF\b|swiss\s*francs?', re.IGNORECASE),
        'SEK': re.compile(r'\bSEK\b|swedish\s*kron', re.IGNORECASE),
        'NOK': re.compile(r'\bNOK\b|norwegian\s*kron', re.IGNORECASE),
        'DKK': re.compile(r'\bDKK\b|danish\s*kron', re.IGNORECASE),
        'PLN': re.compile(r'\bPLN\b|z\u0142|zloty', re.IGNORECASE),
        'CZK': re.compile(r'\bCZK\b|czech\s*korun', re.IGNORECASE),
        'HUF': re.compile(r'\bHUF\b|forints?', re.IGNORECASE),
        'RUB': re.compile(r'\u20bd|\bRUB\b|rubles?', re.IGNORECASE),
        'TRY': re.compile(r'\bTRY\b|\u20ba|turkish\s*lira', re.IGNORECASE),
        'RON': re.compile(r'\bRON\b|romanian\s*lei', re.IGNORECASE),
        'BGN': re.compile(r'\bBGN\b|bulgarian\s*lev', re.IGNORECASE),
        'HRK': re.compile(r'\bHRK\b|kuna', re.IGNORECASE),
        'UAH': re.compile(r'\u20b4|\bUAH\b|hryvnia', re.IGNORECASE),
        # Middle East
        'AED': re.compile(r'\bAED\b|\bdirhams?\b', re.IGNORECASE),
        'SAR': re.compile(r'\bSAR\b|\briyals?\b', re.IGNORECASE),
        'QAR': re.compile(r'\bQAR\b|qatari\s*riyals?', re.IGNORECASE),
        'KWD': re.compile(r'\bKWD\b|kuwaiti\s*dinars?', re.IGNORECASE),
        'BHD': re.compile(r'\bBHD\b|bahraini\s*dinars?', re.IGNORECASE),
        'OMR': re.compile(r'\bOMR\b|omani\s*rials?', re.IGNORECASE),
        'JOD': re.compile(r'\bJOD\b|jordanian\s*dinars?', re.IGNORECASE),
        'ILS': re.compile(r'\u20aa|\bILS\b|shekels?', re.IGNORECASE),
        'IQD': re.compile(r'\bIQD\b|iraqi\s*dinars?', re.IGNORECASE),
        'IRR': re.compile(r'\bIRR\b|iranian\s*rials?', re.IGNORECASE),
        'LBP': re.compile(r'\bLBP\b|lebanese\s*pounds?', re.IGNORECASE),
        # Asia-Pacific
        'JPY': re.compile(r'\u00a5|\bJPY\b|\byen\b', re.IGNORECASE),
        'CNY': re.compile(r'CN\u00a5|\bCNY\b|\bRMB\b|\byuan\b|renminbi', re.IGNORECASE),
        'INR': re.compile(r'\u20b9|\bINR\b|\brupees?\b', re.IGNORECASE),
        'KRW': re.compile(r'\u20a9|\bKRW\b|\bwon\b', re.IGNORECASE),
        'TWD': re.compile(r'\bTWD\b|NT\$|taiwan\s*dollars?', re.IGNORECASE),
        'HKD': re.compile(r'\bHKD\b|HK\$|hong\s*kong\s*dollars?', re.IGNORECASE),
        'SGD': re.compile(r'\bSGD\b|S\$|singapore\s*dollars?', re.IGNORECASE),
        'AUD': re.compile(r'\bAUD\b|A\$|australian\s*dollars?', re.IGNORECASE),
        'NZD': re.compile(r'\bNZD\b|NZ\$|new\s*zealand\s*dollars?', re.IGNORECASE),
        'THB': re.compile(r'\u0e3f|\bTHB\b|\bbaht\b', re.IGNORECASE),
        'MYR': re.compile(r'\bMYR\b|\bRM\b|ringgit', re.IGNORECASE),
        'IDR': re.compile(r'\bIDR\b|indonesian\s*rupiah', re.IGNORECASE),
        'PHP': re.compile(r'\bPHP\b|\u20b1|philippine\s*pesos?', re.IGNORECASE),
        'VND': re.compile(r'\u20ab|\bVND\b|\bdong\b', re.IGNORECASE),
        'PKR': re.compile(r'\bPKR\b|pakistani\s*rupees?', re.IGNORECASE),
        'BDT': re.compile(r'\bBDT\b|\u09f3|taka', re.IGNORECASE),
        'LKR': re.compile(r'\bLKR\b|sri\s*lankan\s*rupees?', re.IGNORECASE),
        'MMK': re.compile(r'\bMMK\b|kyat', re.IGNORECASE),
        'KHR': re.compile(r'\bKHR\b|cambodian\s*riels?', re.IGNORECASE),
        # Africa
        'NGN': re.compile(r'\u20a6|\bNGN\b|\bnaira\b', re.IGNORECASE),
        'GHS': re.compile(r'GH\u20b5|\bGHS\b|\bcedis?\b', re.IGNORECASE),
        'KES': re.compile(r'\bKES\b|KSh|kenyan\s*shillings?', re.IGNORECASE),
        'TZS': re.compile(r'\bTZS\b|tanzanian\s*shillings?', re.IGNORECASE),
        'UGX': re.compile(r'\bUGX\b|ugandan\s*shillings?', re.IGNORECASE),
        'ZAR': re.compile(r'\bZAR\b|\brands?\b', re.IGNORECASE),
        'EGP': re.compile(r'\bEGP\b|egyptian\s*pounds?', re.IGNORECASE),
        'MAD': re.compile(r'\bMAD\b|moroccan\s*dirhams?', re.IGNORECASE),
        'XOF': re.compile(r'\bXOF\b|\bCFA\b|francs?\s*CFA', re.IGNORECASE),
        'XAF': re.compile(r'\bXAF\b|\bFCFA\b', re.IGNORECASE),
        'ETB': re.compile(r'\bETB\b|\bbirr\b', re.IGNORECASE),
        'RWF': re.compile(r'\bRWF\b|rwandan\s*francs?', re.IGNORECASE),
        'ZMW': re.compile(r'\bZMW\b|zambian\s*kwacha', re.IGNORECASE),
        'MWK': re.compile(r'\bMWK\b|malawian\s*kwacha', re.IGNORECASE),
        'BWP': re.compile(r'\bBWP\b|\bpula\b', re.IGNORECASE),
        'MZN': re.compile(r'\bMZN\b|metical|meticais', re.IGNORECASE),
        'AOA': re.compile(r'\bAOA\b|kwanza', re.IGNORECASE),
        'CDF': re.compile(r'\bCDF\b|congolese\s*francs?', re.IGNORECASE),
        'DZD': re.compile(r'\bDZD\b|algerian\s*dinars?', re.IGNORECASE),
        'TND': re.compile(r'\bTND\b|tunisian\s*dinars?', re.IGNORECASE),
        'LYD': re.compile(r'\bLYD\b|libyan\s*dinars?', re.IGNORECASE),
        'SDG': re.compile(r'\bSDG\b|sudanese\s*pounds?', re.IGNORECASE),
    }

    # Pre-compiled explicit currency declaration patterns
    _CODES_PATTERN = '|'.join(ALL_CURRENCY_CODES)
    _EXPLICIT_CURRENCY_PATTERNS = [
        re.compile(rf'\bcurrency\s*[:\-–]\s*({_CODES_PATTERN})\b', re.IGNORECASE),
        re.compile(rf'\b(?:all\s+)?(?:figures|values|amounts)\s+in\s+({_CODES_PATTERN})\b', re.IGNORECASE),
        re.compile(rf'\bin\s+({_CODES_PATTERN})\s*[\'\u2019]?\s*000', re.IGNORECASE),
        re.compile(rf'\(({_CODES_PATTERN})\s*[\'\u2019]?\s*(?:000|mn|m|millions?|billions?|bn)?\)', re.IGNORECASE),
        re.compile(rf'\b({_CODES_PATTERN})\s*[\'\u2019]\s*000', re.IGNORECASE),
    ]
    _CONTROLS_CURRENCY_PATTERN = re.compile(rf'Currency\s+Selection[^A-Z]*({_CODES_PATTERN})\b', re.IGNORECASE)
    _CONTROLS_CONVERSION_PATTERN = re.compile(rf'Conversion\s+Rate[^\d]*([\d,.]+)\s*[^\w]*({_CODES_PATTERN})\s+to\s+({_CODES_PATTERN})', re.IGNORECASE)

    def __init__(self, file_path: Path):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        self.file_type = self.file_path.suffix.lower()
        if self.file_type not in ['.xlsx', '.xls']:
            raise ValueError(f"Unsupported file type: {self.file_type}. Only .xlsx and .xls supported.")

    def parse(self) -> Dict[str, Any]:
        """
        Parse Excel file with enhanced financial analysis.

        Returns:
            Dictionary with sheets, formulas, dependencies, and financial structure
        """
        logger.info(f"Parsing financial Excel: {self.file_path.name}")

        result = {
            "file_name": self.file_path.name,
            "file_path": str(self.file_path.absolute()),
            "file_size": self.file_path.stat().st_size,
            "file_type": self.file_type,
            "sheets": [],
            "sheet_names": [],
            "total_rows": 0,
            "total_columns": 0,
            "model_structure": {
                "has_p_and_l": False,
                "has_balance_sheet": False,
                "has_cash_flow": False,
                "has_saas_metrics": False,
                "has_assumptions": False,
                "has_projections": False,
                "projection_years": 0,
                "historical_years": 0,
                "base_currency": None,
                "sheet_relationships": []
            },
            "all_formulas": {},
            "cross_sheet_references": [],
            "detected_metrics": []
        }

        try:
            # Read file bytes once to avoid multiple disk reads
            file_bytes = self.file_path.read_bytes()

            # Load workbook with formulas (not computed values)
            # Using read_only=True for faster loading on large files
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
            result["sheet_names"] = wb.sheetnames
            logger.debug("Loaded workbook for formulas (read_only=True)")

            # Also load with data_only to get computed values
            # Note: read_only + data_only gives fastest read performance
            wb_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            logger.debug("Loaded workbook for values (read_only=True, data_only=True)")

            # Fallback: if read_only mode can't determine dimensions, reopen without it
            if wb_values.sheetnames:
                test_ws = wb_values[wb_values.sheetnames[0]]
                if test_ws.max_row is None:
                    logger.warning("read_only=True returned None dimensions, falling back to read_only=False")
                    wb.close()
                    wb_values.close()
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
                    result["sheet_names"] = wb.sheetnames
                    wb_values = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)

            # Free the file bytes now that workbooks are loaded
            del file_bytes

            # Pre-read worksheet data in main thread (read_only iterators are NOT thread-safe)
            sheet_data_inputs = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_values = wb_values[sheet_name]
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                # Materialize rows into plain Python lists for thread-safe access
                formula_rows = list(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)) if max_row > 0 and max_col > 0 else []
                value_rows = list(ws_values.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)) if max_row > 0 and max_col > 0 else []
                sheet_data_inputs.append((sheet_name, max_row, max_col, formula_rows, value_rows))

            # Parse sheets in parallel
            workers = min(8, len(sheet_data_inputs))
            sheet_results = {}

            if workers > 1:
                with ThreadPoolExecutor(max_workers=workers) as executor:
                    futures = {
                        executor.submit(self._parse_sheet_from_rows, name, max_r, max_c, f_rows, v_rows): name
                        for name, max_r, max_c, f_rows, v_rows in sheet_data_inputs
                    }
                    for future in as_completed(futures):
                        name = futures[future]
                        try:
                            sheet_results[name] = future.result()
                        except Exception as e:
                            logger.warning(f"Failed to parse sheet '{name}': {e}")
                            sheet_results[name] = None
            else:
                for name, max_r, max_c, f_rows, v_rows in sheet_data_inputs:
                    sheet_results[name] = self._parse_sheet_from_rows(name, max_r, max_c, f_rows, v_rows)

            # Reassemble results in original sheet order (must be sequential — updates shared state)
            for sheet_name in wb.sheetnames:
                sheet_data = sheet_results.get(sheet_name)
                if sheet_data and sheet_data["rows"] > 0:
                    result["sheets"].append(sheet_data)
                    result["total_rows"] += sheet_data["rows"]
                    result["total_columns"] = max(result["total_columns"], sheet_data["columns"])

                    # Update model structure based on sheet type
                    self._update_model_structure(result["model_structure"], sheet_data)

                    # Collect all formulas
                    for cell, formula in sheet_data["formulas"].items():
                        result["all_formulas"][f"{sheet_name}!{cell}"] = formula

            # Analyze cross-sheet references
            result["cross_sheet_references"] = self._find_cross_sheet_references(result["all_formulas"])

            # Build sheet relationships
            result["model_structure"]["sheet_relationships"] = self._build_sheet_relationships(
                result["cross_sheet_references"]
            )

            # Detect time periods
            self._detect_time_periods(result)

            # Detect currency: explicit declarations > Controls sheet > regex fallback
            currency_settings = self._detect_currency_from_controls(result["sheets"])
            result["model_structure"]["currency_settings"] = currency_settings

            explicit_currency = self._detect_explicit_currency(result["sheets"])
            base_currency = self._detect_currency(result["sheets"])
            # Priority: explicit declaration on any sheet > Controls sheet > regex counting
            if explicit_currency:
                base_currency = explicit_currency
            if currency_settings and currency_settings.get("display_currency"):
                base_currency = currency_settings["display_currency"]
            result["model_structure"]["base_currency"] = base_currency

            result["model_structure"]["per_sheet_currencies"] = self._assign_per_sheet_currencies(
                result["sheets"], currency_settings, base_currency,
                has_explicit_currency=explicit_currency is not None
            )

            # Close workbooks to release file handles
            wb.close()
            wb_values.close()

            logger.success(
                f"Parsed {len(result['sheets'])} sheet(s), "
                f"{len(result['all_formulas'])} formulas, "
                f"{len(result['cross_sheet_references'])} cross-sheet refs"
            )

            return result

        except Exception as e:
            logger.error(f"Failed to parse financial Excel: {e}")
            result["error"] = str(e)
            # Try to close workbooks if they were opened
            try:
                if 'wb' in locals():
                    wb.close()
                if 'wb_values' in locals():
                    wb_values.close()
            except Exception:
                pass
            return result

    def _parse_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        ws_values: openpyxl.worksheet.worksheet.Worksheet,
        sheet_name: str
    ) -> Dict[str, Any]:
        """Parse a single sheet with formula and structure extraction."""

        # Get dimensions
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            return {
                "sheet_name": sheet_name,
                "sheet_type": "empty",
                "rows": 0,
                "columns": 0,
                "headers": [],
                "row_labels": [],
                "data": [],
                "formulas": {},
                "cell_dependencies": {},
                "time_periods": [],
                "financial_metrics": {}
            }

        # Extract headers (first row with content) - using iter_rows for performance
        headers = []
        header_row = 1
        for row_idx, row in enumerate(ws_values.iter_rows(min_row=1, max_row=min(10, max_row), min_col=1, max_col=max_col), start=1):
            row_has_content = any(cell.value for cell in row)
            if row_has_content:
                header_row = row_idx
                break

        # Extract header values from the identified header row
        for row in ws_values.iter_rows(min_row=header_row, max_row=header_row, min_col=1, max_col=max_col):
            for col_idx, cell in enumerate(row, start=1):
                headers.append(str(cell.value) if cell.value else f"Col{col_idx}")

        # Extract row labels and data (values only) - using iter_rows for performance
        row_labels = []
        data = []
        for row in ws_values.iter_rows(min_row=header_row + 1, max_row=max_row, min_col=1, max_col=max_col):
            row_data = []
            first_cell_value = None
            for col_idx, cell in enumerate(row):
                value = cell.value
                # Capture first column as row label
                if col_idx == 0:
                    first_cell_value = value
                # Handle numeric formatting
                if isinstance(value, (int, float)):
                    row_data.append(value)
                elif value is None:
                    row_data.append(None)
                else:
                    row_data.append(str(value))
            row_labels.append(str(first_cell_value) if first_cell_value else "")
            data.append(row_data)

        # Extract formulas (with cell limit to prevent hanging on large sparse sheets)
        formulas = {}
        cell_dependencies = {}
        total_cells = max_row * max_col
        formula_scan_skipped = False

        if total_cells <= self.MAX_CELLS_FOR_FORMULA_SCAN:
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        coord = cell.coordinate
                        formula = cell.value
                        formulas[coord] = formula

                        # Parse dependencies
                        deps = self._parse_formula_dependencies(formula)
                        if deps:
                            cell_dependencies[coord] = deps
        else:
            # Sheet is too large - skip formula extraction to prevent hanging
            logger.warning(
                f"Sheet '{sheet_name}' has {total_cells:,} cells (max: {self.MAX_CELLS_FOR_FORMULA_SCAN:,}). "
                f"Skipping formula extraction to prevent slow processing."
            )
            formula_scan_skipped = True

        # Detect sheet type
        sheet_type = self._detect_sheet_type(row_labels, headers)

        # Detect time periods from headers
        time_periods = self._extract_time_periods(headers)

        # Extract financial metrics from this sheet
        financial_metrics = self._extract_sheet_metrics(row_labels, data, headers)

        return {
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
            "rows": max_row - header_row,
            "columns": max_col,
            "headers": headers,
            "row_labels": row_labels,
            "data": data,
            "formulas": formulas,
            "cell_dependencies": cell_dependencies,
            "time_periods": time_periods,
            "financial_metrics": financial_metrics,
            "formula_count": len(formulas),
            "header_row": header_row,
            "formula_scan_skipped": formula_scan_skipped
        }

    def _parse_sheet_from_rows(
        self,
        sheet_name: str,
        max_row: int,
        max_col: int,
        formula_rows: list,
        value_rows: list,
    ) -> Dict[str, Any]:
        """Parse a single sheet from pre-materialized row data (thread-safe)."""

        if max_row == 0 or max_col == 0 or not value_rows:
            return {
                "sheet_name": sheet_name,
                "sheet_type": "empty",
                "rows": 0,
                "columns": 0,
                "headers": [],
                "row_labels": [],
                "data": [],
                "formulas": {},
                "cell_dependencies": {},
                "time_periods": [],
                "financial_metrics": {}
            }

        # Extract headers (first row with content)
        headers = []
        header_row = 1
        for row_idx, row in enumerate(value_rows[:min(10, len(value_rows))], start=1):
            row_has_content = any(cell.value for cell in row)
            if row_has_content:
                header_row = row_idx
                break

        # Extract header values from the identified header row
        header_row_cells = value_rows[header_row - 1] if header_row <= len(value_rows) else []
        for col_idx, cell in enumerate(header_row_cells, start=1):
            headers.append(str(cell.value) if cell.value else f"Col{col_idx}")

        # Extract row labels and data (values only)
        row_labels = []
        data = []
        for row in value_rows[header_row:]:
            row_data = []
            first_cell_value = None
            for col_idx, cell in enumerate(row):
                value = cell.value
                if col_idx == 0:
                    first_cell_value = value
                if isinstance(value, (int, float)):
                    row_data.append(value)
                elif value is None:
                    row_data.append(None)
                else:
                    row_data.append(str(value))
            row_labels.append(str(first_cell_value) if first_cell_value else "")
            data.append(row_data)

        # Extract formulas (with cell limit to prevent hanging on large sparse sheets)
        formulas = {}
        cell_dependencies = {}
        total_cells = max_row * max_col
        formula_scan_skipped = False

        if total_cells <= self.MAX_CELLS_FOR_FORMULA_SCAN:
            for row in formula_rows:
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        coord = cell.coordinate
                        formula = cell.value
                        formulas[coord] = formula

                        deps = self._parse_formula_dependencies(formula)
                        if deps:
                            cell_dependencies[coord] = deps
        else:
            logger.warning(
                f"Sheet '{sheet_name}' has {total_cells:,} cells (max: {self.MAX_CELLS_FOR_FORMULA_SCAN:,}). "
                f"Skipping formula extraction to prevent slow processing."
            )
            formula_scan_skipped = True

        # Detect sheet type
        sheet_type = self._detect_sheet_type(row_labels, headers)

        # Detect time periods from headers
        time_periods = self._extract_time_periods(headers)

        # Extract financial metrics from this sheet
        financial_metrics = self._extract_sheet_metrics(row_labels, data, headers)

        return {
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
            "rows": max_row - header_row,
            "columns": max_col,
            "headers": headers,
            "row_labels": row_labels,
            "data": data,
            "formulas": formulas,
            "cell_dependencies": cell_dependencies,
            "time_periods": time_periods,
            "financial_metrics": financial_metrics,
            "formula_count": len(formulas),
            "header_row": header_row,
            "formula_scan_skipped": formula_scan_skipped
        }

    def _parse_formula_dependencies(self, formula: str) -> List[str]:
        """Extract cell references from a formula."""
        # Pattern to match cell references like A1, $B$2, Sheet1!A1, 'Sheet Name'!A1
        # Simple cell reference: A1, $A$1, A$1, $A1
        simple_pattern = r'\$?[A-Z]+\$?\d+'
        # Sheet reference: Sheet1!A1, 'Sheet Name'!A1
        sheet_pattern = r"(?:'[^']+'|[A-Za-z0-9_]+)!\$?[A-Z]+\$?\d+"

        refs = []

        # Find sheet references first
        sheet_refs = re.findall(sheet_pattern, formula)
        refs.extend(sheet_refs)

        # Remove sheet references from formula to find simple refs
        temp_formula = formula
        for ref in sheet_refs:
            temp_formula = temp_formula.replace(ref, '')

        # Find simple cell references
        simple_refs = re.findall(simple_pattern, temp_formula)
        refs.extend(simple_refs)

        return list(set(refs))  # Remove duplicates

    def _detect_sheet_type(self, row_labels: List[str], headers: List[str]) -> str:
        """Detect the type of financial statement from content."""
        all_text = ' '.join(row_labels + headers).lower()

        # Count keyword matches for each type
        pl_matches = sum(1 for kw in self.PL_KEYWORDS if kw in all_text)
        bs_matches = sum(1 for kw in self.BS_KEYWORDS if kw in all_text)
        cf_matches = sum(1 for kw in self.CF_KEYWORDS if kw in all_text)
        saas_matches = sum(1 for kw in self.SAAS_KEYWORDS if kw in all_text)
        assumptions_matches = sum(1 for kw in self.ASSUMPTIONS_KEYWORDS if kw in all_text)

        # Return the type with most matches (minimum 2 matches required)
        type_scores = {
            'p_and_l': pl_matches,
            'balance_sheet': bs_matches,
            'cash_flow': cf_matches,
            'saas_metrics': saas_matches,
            'assumptions': assumptions_matches
        }

        best_type = max(type_scores, key=type_scores.get)
        if type_scores[best_type] >= 2:
            return best_type

        return 'other'

    def _extract_time_periods(self, headers: List[str]) -> List[Dict[str, Any]]:
        """Extract time periods from headers."""
        time_periods = []

        # Patterns to match
        year_pattern = r'(20\d{2}|19\d{2})'
        quarter_pattern = r'Q([1-4])\s*(20\d{2}|19\d{2})?'
        month_pattern = r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s-]*(20\d{2}|19\d{2})?'
        fy_pattern = r'FY\s*(20\d{2}|19\d{2}|\d{2})'
        projection_pattern = r'(forecast|projected|budget|plan)'

        for i, header in enumerate(headers):
            if not header:
                continue

            header_str = str(header)
            period_info = {
                "column_index": i,
                "header": header_str,
                "period_type": None,
                "year": None,
                "is_projection": False
            }

            # Check if it's a projection
            if re.search(projection_pattern, header_str, re.IGNORECASE):
                period_info["is_projection"] = True

            # Try to match year
            year_match = re.search(year_pattern, header_str)
            if year_match:
                period_info["year"] = int(year_match.group(1))
                period_info["period_type"] = "year"

            # Try to match quarter
            quarter_match = re.search(quarter_pattern, header_str, re.IGNORECASE)
            if quarter_match:
                period_info["period_type"] = "quarter"
                period_info["quarter"] = int(quarter_match.group(1))
                if quarter_match.group(2):
                    period_info["year"] = int(quarter_match.group(2))

            # Try to match month
            month_match = re.search(month_pattern, header_str, re.IGNORECASE)
            if month_match:
                period_info["period_type"] = "month"
                period_info["month"] = month_match.group(1)
                if month_match.group(2):
                    period_info["year"] = int(month_match.group(2))

            # Try FY pattern
            fy_match = re.search(fy_pattern, header_str, re.IGNORECASE)
            if fy_match:
                period_info["period_type"] = "fiscal_year"
                year_str = fy_match.group(1)
                if len(year_str) == 2:
                    period_info["year"] = 2000 + int(year_str)
                else:
                    period_info["year"] = int(year_str)

            if period_info["period_type"]:
                time_periods.append(period_info)

        return time_periods

    def _extract_sheet_metrics(
        self,
        row_labels: List[str],
        data: List[List[Any]],
        headers: List[str]
    ) -> Dict[str, Any]:
        """Extract financial metrics found in this sheet."""
        metrics = {}

        # Metric patterns to look for
        metric_patterns = {
            'revenue': r'^(total\s+)?revenue$|^sales$|^net\s+sales$',
            'gross_profit': r'^gross\s+profit$|^gross\s+margin$',
            'operating_income': r'^operating\s+(income|profit)$|^ebit$',
            'ebitda': r'^ebitda$',
            'net_income': r'^net\s+(income|profit)$',
            'total_assets': r'^total\s+assets$',
            'total_liabilities': r'^total\s+liabilities$',
            'cash': r'^cash(\s+and\s+cash\s+equivalents)?$',
            'arr': r'^arr$|^annual\s+recurring\s+revenue$',
            'mrr': r'^mrr$|^monthly\s+recurring\s+revenue$',
            'customers': r'^(total\s+)?customers?$|^customer\s+count$',
            'churn': r'^churn(\s+rate)?$',
            'cac': r'^cac$|^customer\s+acquisition\s+cost$',
            'ltv': r'^ltv$|^lifetime\s+value$|^clv$'
        }

        for i, label in enumerate(row_labels):
            if not label:
                continue

            label_lower = label.lower().strip()

            for metric_name, pattern in metric_patterns.items():
                if re.match(pattern, label_lower, re.IGNORECASE):
                    # Get values for this row
                    if i < len(data):
                        values = []
                        for j, val in enumerate(data[i]):
                            if isinstance(val, (int, float)) and val is not None:
                                values.append({
                                    "column": j,
                                    "header": headers[j] if j < len(headers) else f"Col{j}",
                                    "value": val
                                })

                        if values:
                            metrics[metric_name] = {
                                "row_label": label,
                                "row_index": i,
                                "values": values
                            }
                    break

        return metrics

    def _update_model_structure(
        self,
        model_structure: Dict[str, Any],
        sheet_data: Dict[str, Any]
    ):
        """Update model structure based on sheet analysis."""
        sheet_type = sheet_data["sheet_type"]

        if sheet_type == 'p_and_l':
            model_structure["has_p_and_l"] = True
        elif sheet_type == 'balance_sheet':
            model_structure["has_balance_sheet"] = True
        elif sheet_type == 'cash_flow':
            model_structure["has_cash_flow"] = True
        elif sheet_type == 'saas_metrics':
            model_structure["has_saas_metrics"] = True
        elif sheet_type == 'assumptions':
            model_structure["has_assumptions"] = True

    def _find_cross_sheet_references(self, all_formulas: Dict[str, str]) -> List[Dict[str, str]]:
        """Find formulas that reference other sheets."""
        cross_refs = []

        # Pattern for sheet references
        sheet_ref_pattern = r"(?:'([^']+)'|([A-Za-z0-9_]+))!(\$?[A-Z]+\$?\d+)"

        for cell_ref, formula in all_formulas.items():
            source_sheet = cell_ref.split('!')[0]

            matches = re.findall(sheet_ref_pattern, formula)
            for match in matches:
                target_sheet = match[0] if match[0] else match[1]
                target_cell = match[2]

                if target_sheet != source_sheet:
                    cross_refs.append({
                        "source": cell_ref,
                        "target_sheet": target_sheet,
                        "target_cell": target_cell,
                        "formula": formula
                    })

        return cross_refs

    def _build_sheet_relationships(
        self,
        cross_refs: List[Dict[str, str]]
    ) -> List[Tuple[str, str]]:
        """Build list of sheet-to-sheet relationships."""
        relationships = set()

        for ref in cross_refs:
            source_sheet = ref["source"].split('!')[0]
            target_sheet = ref["target_sheet"]

            # Store as tuple (source, target) - source depends on target
            relationships.add((source_sheet, target_sheet))

        return list(relationships)

    def _detect_time_periods(self, result: Dict[str, Any]):
        """Analyze time periods across all sheets to determine projections."""
        import datetime
        current_year = datetime.datetime.now().year

        all_years = set()
        has_projections = False

        for sheet in result["sheets"]:
            for period in sheet.get("time_periods", []):
                year = period.get("year")
                if year:
                    all_years.add(year)
                    if period.get("is_projection") or year > current_year:
                        has_projections = True

        if all_years:
            historical_years = [y for y in all_years if y <= current_year]
            projection_years = [y for y in all_years if y > current_year]

            result["model_structure"]["historical_years"] = len(historical_years)
            result["model_structure"]["projection_years"] = len(projection_years)
            result["model_structure"]["has_projections"] = has_projections or len(projection_years) > 0

    def _detect_currency(self, sheets: List[Dict[str, Any]]) -> Optional[str]:
        """Try to detect the currency used in the model by counting pattern matches across all sheets."""
        match_counts: Dict[str, int] = {c: 0 for c in self._CURRENCY_PATTERNS}

        for sheet in sheets:
            all_text = ' '.join(
                sheet.get("headers", []) +
                sheet.get("row_labels", []) +
                [str(v) for row in sheet.get("data", []) for v in row if isinstance(v, str)]
            )

            for currency, compiled_re in self._CURRENCY_PATTERNS.items():
                matches = compiled_re.findall(all_text)
                match_counts[currency] += len(matches)

            # Early exit: if one currency has 3+ matches and no other has any, skip remaining sheets
            non_zero = [(c, n) for c, n in match_counts.items() if n > 0]
            if len(non_zero) == 1 and non_zero[0][1] >= 3:
                return non_zero[0][0]

        best = max(match_counts, key=match_counts.get)
        if match_counts[best] > 0:
            return best

        return None

    def _detect_explicit_currency(self, sheets: List[Dict[str, Any]]) -> Optional[str]:
        """
        Scan all sheets for explicit currency declarations like
        "Currency: EGP", "in EGP '000", "Figures in EGP", "(EGP)", etc.
        Prioritizes the first sheet (cover/summary pages often state the currency).
        Returns the currency code if found, else None.
        """
        for sheet in sheets:
            all_text = ' '.join(
                sheet.get("headers", []) +
                sheet.get("row_labels", []) +
                [str(v) for row in sheet.get("data", []) for v in row if isinstance(v, str)]
            )
            for compiled_re in self._EXPLICIT_CURRENCY_PATTERNS:
                match = compiled_re.search(all_text)
                if match:
                    currency = match.group(1).upper()
                    logger.info(f"Explicit currency declaration found on sheet '{sheet.get('sheet_name')}': {currency}")
                    return currency

        return None

    def _detect_currency_from_controls(self, sheets: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        """
        Extract currency settings from a Controls/Settings sheet.
        Looks for 'Currency Selection' and 'Conversion Rate' patterns.
        Returns dict with display_currency, local_currency, conversion_rate, or None.
        """
        for sheet in sheets:
            sheet_name_lower = sheet.get("sheet_name", "").lower()
            if not any(kw in sheet_name_lower for kw in ("control", "setting", "config", "assumption", "input")):
                continue

            all_text = ' '.join(
                sheet.get("headers", []) +
                sheet.get("row_labels", []) +
                [str(v) for row in sheet.get("data", []) for v in row if v is not None]
            )

            cs_match = self._CONTROLS_CURRENCY_PATTERN.search(all_text)
            cr_match = self._CONTROLS_CONVERSION_PATTERN.search(all_text)

            if cs_match or cr_match:
                result = {"display_currency": None, "local_currency": None, "conversion_rate": None, "has_conversion": False}

                if cs_match:
                    result["display_currency"] = cs_match.group(1).upper()

                if cr_match:
                    try:
                        result["conversion_rate"] = float(cr_match.group(1).replace(',', ''))
                    except ValueError:
                        pass
                    from_currency = cr_match.group(2).upper()
                    to_currency = cr_match.group(3).upper()
                    result["local_currency"] = from_currency
                    result["has_conversion"] = True
                    # If display_currency wasn't found, infer from conversion target
                    if not result["display_currency"]:
                        result["display_currency"] = to_currency

                logger.info(f"Currency settings from Controls sheet: {result}")
                return result

        return None

    def _assign_per_sheet_currencies(self, sheets: List[Dict[str, Any]], currency_settings: Optional[Dict[str, Any]], base_currency: Optional[str], has_explicit_currency: bool = False) -> Dict[str, Optional[str]]:
        """
        Assign currency to each sheet based on Controls settings and sheet name patterns.
        Historical sheets get local_currency, projection sheets get display_currency.
        When has_explicit_currency is True, skip unreliable per-sheet regex detection.
        """
        per_sheet = {}

        for sheet in sheets:
            name = sheet.get("sheet_name", "")
            name_lower = name.lower()

            if currency_settings and currency_settings.get("has_conversion"):
                local = currency_settings.get("local_currency")
                display = currency_settings.get("display_currency")

                # Historical sheets → local currency
                if any(kw in name_lower for kw in ("hist", "historical", "actual")):
                    per_sheet[name] = local
                # Controls/settings sheets → no specific currency
                elif any(kw in name_lower for kw in ("control", "setting", "config", "input", "assumption")):
                    per_sheet[name] = display
                else:
                    # Projection/summary/highlight sheets → display currency
                    per_sheet[name] = display
            else:
                # Use base_currency directly when we have a high-confidence explicit
                # declaration. Per-sheet regex is unreliable (e.g. "expense" → PEN).
                per_sheet[name] = base_currency

        return per_sheet


def parse_excel_financial(file_path: str) -> Dict[str, Any]:
    """
    Convenience function to parse Excel file for financial analysis.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with parsed content including formulas and structure

    Example:
        >>> result = parse_excel_financial("model.xlsx")
        >>> print(f"Sheets: {result['sheet_names']}")
        >>> print(f"Has P&L: {result['model_structure']['has_p_and_l']}")
    """
    parser = FinancialExcelParser(file_path)
    return parser.parse()


def main():
    """CLI interface for testing."""
    import argparse

    parser = argparse.ArgumentParser(description="Parse Excel files for financial analysis")
    parser.add_argument("file", help="Path to Excel file")
    parser.add_argument("--output", help="Output JSON file path")
    parser.add_argument("--verbose", action="store_true", help="Verbose logging")

    args = parser.parse_args()

    if args.verbose:
        logger.remove()
        logger.add(sys.stderr, level="DEBUG")

    # Parse file
    result = parse_excel_financial(args.file)

    # Output
    if args.output:
        output_path = Path(args.output)
        output_path.write_text(json.dumps(result, indent=2, default=str))
        logger.info(f"Saved results to {output_path}")
    else:
        print(f"\n{'='*60}")
        print(f"Financial Excel Analysis: {result['file_name']}")
        print(f"{'='*60}\n")

        print("Model Structure:")
        structure = result.get('model_structure', {})
        print(f"  - P&L Sheet: {structure.get('has_p_and_l', False)}")
        print(f"  - Balance Sheet: {structure.get('has_balance_sheet', False)}")
        print(f"  - Cash Flow: {structure.get('has_cash_flow', False)}")
        print(f"  - SaaS Metrics: {structure.get('has_saas_metrics', False)}")
        print(f"  - Assumptions: {structure.get('has_assumptions', False)}")
        print(f"  - Historical Years: {structure.get('historical_years', 0)}")
        print(f"  - Projection Years: {structure.get('projection_years', 0)}")
        print(f"  - Currency: {structure.get('base_currency', 'Unknown')}")

        print(f"\nSheets ({len(result['sheets'])}):")
        for sheet in result['sheets']:
            print(f"  - {sheet['sheet_name']} ({sheet['sheet_type']}): "
                  f"{sheet['rows']} rows, {sheet['formula_count']} formulas")

        print(f"\nFormulas: {len(result.get('all_formulas', {}))}")
        print(f"Cross-sheet References: {len(result.get('cross_sheet_references', []))}")

        if result.get('error'):
            print(f"\nError: {result['error']}")

        print(f"\n{'='*60}\n")


if __name__ == "__main__":
    main()
