"""
Excel and CSV parsing tool for financial models and data extraction.
Handles multiple sheets, formulas, and preserves structure.
"""

import sys
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Any, Optional
import json
from loguru import logger

try:
    import pandas as pd
    import openpyxl
except ImportError:
    logger.error("Required packages not installed. Run: pip install pandas openpyxl")
    sys.exit(1)


# Patterns that indicate a header row
HEADER_PATTERNS = [
    r'\b(account|description|category|item|name|type|total|amount)\b',  # Common header terms
    r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b',  # Month abbreviations
    r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\b',
    r'\b(q1|q2|q3|q4|quarter)\b',  # Quarters
    r'\b(ytd|year.to.date|fy\d{2,4}|20\d{2}|19\d{2})\b',  # Year references
    r'\b(budget|actual|forecast|variance|prior|current)\b',  # Financial terms
]
HEADER_REGEX = re.compile('|'.join(HEADER_PATTERNS), re.IGNORECASE)


class ExcelParser:
    """Excel/CSV parser with formula and structure preservation."""

    MAX_ROWS_PER_SHEET = 50_000  # Cap for very large data sheets (e.g. loan books, transaction logs)

    def __init__(self, file_path: Path):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        self.file_type = self.file_path.suffix.lower()

    def parse(self) -> Dict[str, Any]:
        """
        Parse Excel or CSV file.

        Returns:
            Dictionary with all sheets and metadata
        """
        logger.info(f"Parsing Excel file: {self.file_path.name}")

        result = {
            "file_name": self.file_path.name,
            "file_path": str(self.file_path.absolute()),
            "file_size": self.file_path.stat().st_size,
            "file_type": self.file_type,
            "sheets": [],
            "sheet_names": [],
            "total_rows": 0,
            "total_columns": 0,
            "has_formulas": False,
            "metadata": {}
        }

        try:
            if self.file_type == '.csv':
                result = self._parse_csv()
            else:
                result = self._parse_excel()

            logger.success(f"Parsed {len(result['sheets'])} sheet(s), {result['total_rows']} total rows")
            return result

        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            result["error"] = str(e)
            return result

    def _parse_csv(self) -> Dict[str, Any]:
        """Parse CSV file."""
        result = {
            "file_name": self.file_path.name,
            "file_path": str(self.file_path.absolute()),
            "file_size": self.file_path.stat().st_size,
            "file_type": ".csv",
            "sheets": [],
            "sheet_names": ["Sheet1"],
            "total_rows": 0,
            "total_columns": 0,
            "has_formulas": False,
            "metadata": {}
        }

        # Read CSV
        df = pd.read_csv(self.file_path)
        # Trim to actual data boundaries (removes empty trailing rows/columns)
        df = self._trim_to_data_boundaries(df)

        truncated = False
        if len(df) > self.MAX_ROWS_PER_SHEET:
            logger.warning(f"CSV has {len(df):,} rows, truncating to {self.MAX_ROWS_PER_SHEET:,}")
            df = df.head(self.MAX_ROWS_PER_SHEET)
            truncated = True

        sheet_data = {
            "sheet_name": "Sheet1",
            "rows": len(df),
            "columns": len(df.columns),
            "headers": df.columns.tolist(),
            "data": df.values.tolist(),
            "data_frame": df.to_dict('records'),  # More structured format
            "statistics": self._calculate_statistics(df),
            "truncated": truncated,
        }

        result["sheets"].append(sheet_data)
        result["total_rows"] = len(df)
        result["total_columns"] = len(df.columns)

        return result

    def _parse_excel(self) -> Dict[str, Any]:
        """Parse Excel file (.xlsx, .xls)."""
        result = {
            "file_name": self.file_path.name,
            "file_path": str(self.file_path.absolute()),
            "file_size": self.file_path.stat().st_size,
            "file_type": self.file_type,
            "sheets": [],
            "sheet_names": [],
            "total_rows": 0,
            "total_columns": 0,
            "has_formulas": False,
            "metadata": {}
        }

        # Read all sheets
        try:
            excel_file = pd.ExcelFile(self.file_path)
            result["sheet_names"] = excel_file.sheet_names

            # Load openpyxl workbook once for formula detection (xlsx only)
            openpyxl_wb = None
            if self.file_type == '.xlsx':
                try:
                    openpyxl_wb = openpyxl.load_workbook(self.file_path, data_only=False, read_only=True)
                    logger.debug("Loaded workbook for formula detection (read-only)")
                except Exception as e:
                    logger.warning(f"Could not load workbook with openpyxl: {e}")

            def _process_sheet(sheet_name):
                """Process a single sheet — suitable for parallel execution."""
                logger.info(f"Processing sheet: '{sheet_name}'")
                try:
                    # Read sheet once with no header, then detect header from in-memory data
                    df_raw = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
                    header_row = self._detect_header_row_from_df(df_raw)
                    if header_row > 0:
                        df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
                        df.columns = [str(c) if pd.notna(c) else f"Unnamed: {i}" for i, c in enumerate(df_raw.iloc[header_row])]
                    else:
                        df = df_raw.iloc[1:].reset_index(drop=True)
                        df.columns = [str(c) if pd.notna(c) else f"Unnamed: {i}" for i, c in enumerate(df_raw.iloc[0])]
                    df = self._trim_to_data_boundaries(df)
                except Exception as e:
                    logger.warning(f"Failed to read sheet '{sheet_name}': {e}")
                    return None

                if df.empty:
                    logger.info(f"  Skipping sheet '{sheet_name}' (empty)")
                    return None

                truncated = False
                if len(df) > self.MAX_ROWS_PER_SHEET:
                    logger.warning(f"Sheet '{sheet_name}' has {len(df):,} rows, truncating to {self.MAX_ROWS_PER_SHEET:,}")
                    df = df.head(self.MAX_ROWS_PER_SHEET)
                    truncated = True

                sheet_data = {
                    "sheet_name": sheet_name,
                    "rows": len(df),
                    "columns": len(df.columns),
                    "headers": df.columns.tolist(),
                    "data": df.values.tolist(),
                    "data_frame": df.to_dict('records'),
                    "statistics": self._calculate_statistics(df),
                    "has_formulas": False,
                    "charts": [],
                    "truncated": truncated,
                }

                if openpyxl_wb is not None:
                    try:
                        sheet_data["has_formulas"] = self._detect_formulas_from_wb(openpyxl_wb, sheet_name)
                    except Exception as e:
                        logger.debug(f"Formula detection failed for {sheet_name}: {e}")

                logger.info(f"  Parsed '{sheet_name}': {len(df)} rows, {len(df.columns)} cols")
                return sheet_data

            # Process sheets in parallel
            sheet_names = excel_file.sheet_names
            workers = min(8, len(sheet_names))

            if workers > 1:
                with ThreadPoolExecutor(max_workers=workers) as executor:
                    futures = {executor.submit(_process_sheet, name): name for name in sheet_names}
                    # Collect in original order
                    sheet_results = {}
                    for future in as_completed(futures):
                        name = futures[future]
                        sheet_results[name] = future.result()
                    for name in sheet_names:
                        sheet_data = sheet_results.get(name)
                        if sheet_data:
                            result["sheets"].append(sheet_data)
                            result["total_rows"] += sheet_data["rows"]
                            result["total_columns"] = max(result["total_columns"], sheet_data["columns"])
                            if sheet_data.get("has_formulas"):
                                result["has_formulas"] = True
            else:
                for name in sheet_names:
                    sheet_data = _process_sheet(name)
                    if sheet_data:
                        result["sheets"].append(sheet_data)
                        result["total_rows"] += sheet_data["rows"]
                        result["total_columns"] = max(result["total_columns"], sheet_data["columns"])
                        if sheet_data.get("has_formulas"):
                            result["has_formulas"] = True

            if openpyxl_wb is not None:
                openpyxl_wb.close()

        except Exception as e:
            logger.error(f"Error parsing Excel sheets: {e}")
            result["error"] = str(e)

        return result

    def _extract_charts(self, wb: openpyxl.Workbook, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract chart metadata from a worksheet."""
        charts = []
        try:
            ws = wb[sheet_name]
            # Access charts through _charts attribute
            if hasattr(ws, '_charts') and ws._charts:
                for chart in ws._charts:
                    chart_info = {
                        "type": chart.__class__.__name__.replace('Chart', ''),
                        "title": None,
                        "x_axis": None,
                        "y_axis": None,
                    }
                    # Extract chart title
                    if hasattr(chart, 'title') and chart.title:
                        if hasattr(chart.title, 'text'):
                            chart_info["title"] = chart.title.text
                        else:
                            chart_info["title"] = str(chart.title)
                    # Extract axis labels
                    if hasattr(chart, 'x_axis') and chart.x_axis and hasattr(chart.x_axis, 'title') and chart.x_axis.title:
                        if hasattr(chart.x_axis.title, 'text'):
                            chart_info["x_axis"] = chart.x_axis.title.text
                        else:
                            chart_info["x_axis"] = str(chart.x_axis.title)
                    if hasattr(chart, 'y_axis') and chart.y_axis and hasattr(chart.y_axis, 'title') and chart.y_axis.title:
                        if hasattr(chart.y_axis.title, 'text'):
                            chart_info["y_axis"] = chart.y_axis.title.text
                        else:
                            chart_info["y_axis"] = str(chart.y_axis.title)
                    charts.append(chart_info)
        except Exception as e:
            logger.debug(f"Could not extract charts from {sheet_name}: {e}")
        return charts

    def _detect_formulas_from_wb(self, wb: openpyxl.Workbook, sheet_name: str) -> bool:
        """Detect if sheet contains formulas using pre-loaded workbook."""
        try:
            ws = wb[sheet_name]

            # Check first 100 cells for formulas
            check_count = 0
            for row in ws.iter_rows(max_row=20, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        return True
                    check_count += 1
                    if check_count > 100:
                        return False

            return False
        except Exception as e:
            logger.debug(f"Formula detection error: {e}")
            return False

    def _detect_formulas(self, sheet_name: str) -> bool:
        """Detect if sheet contains formulas using openpyxl (legacy method)."""
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=False, read_only=True)
            result = self._detect_formulas_from_wb(wb, sheet_name)
            wb.close()
            return result
        except Exception as e:
            logger.debug(f"Formula detection error: {e}")
            return False

    def _calculate_statistics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate basic statistics for numeric columns."""
        stats = {
            "numeric_columns": [],
            "text_columns": [],
            "date_columns": [],
            "null_counts": {}
        }

        for col in df.columns:
            # Check data type
            if pd.api.types.is_numeric_dtype(df[col]):
                stats["numeric_columns"].append(col)
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                stats["date_columns"].append(col)
            else:
                stats["text_columns"].append(col)

            # Null count
            null_count = df[col].isnull().sum()
            if null_count > 0:
                stats["null_counts"][col] = int(null_count)

        return stats

    def _trim_to_data_boundaries(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Trim DataFrame to actual data boundaries, removing trailing empty rows/columns.

        Essential for sparse Excel files where only a small percentage of allocated
        cells contain actual data (e.g., 1000 rows allocated but only 50 used).

        Args:
            df: DataFrame from pandas.read_excel

        Returns:
            Trimmed DataFrame with empty trailing rows/columns removed
        """
        if df.empty:
            return df

        original_shape = df.shape

        # Drop rows where ALL values are NaN
        df_trimmed = df.dropna(how='all')

        # Drop columns where ALL values are NaN
        df_trimmed = df_trimmed.dropna(axis=1, how='all')

        if df_trimmed.shape != original_shape:
            logger.debug(f"Trimmed from {original_shape} to {df_trimmed.shape}")

        return df_trimmed

    def _detect_header_row(self, sheet_name: str) -> int:
        """
        Detect the actual header row in an Excel sheet.

        Financial spreadsheets often have title rows (company name, report title, date)
        before the actual data headers. This method scans the first 10 rows to find
        the row that looks most like a header row.

        Args:
            sheet_name: Name of the sheet to analyze

        Returns:
            0-indexed row number to use as header (for pandas header parameter)
        """
        try:
            # Read first 15 rows without headers to analyze
            df_preview = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=None,
                nrows=15
            )

            if df_preview.empty:
                return 0

            best_row = 0
            best_score = 0

            for row_idx in range(min(10, len(df_preview))):
                row = df_preview.iloc[row_idx]
                score = self._score_header_row(row)

                if score > best_score:
                    best_score = score
                    best_row = row_idx

            if best_row > 0:
                logger.debug(f"Detected header row {best_row + 1} for sheet '{sheet_name}' (score: {best_score})")

            return best_row

        except Exception as e:
            logger.debug(f"Header detection failed for '{sheet_name}': {e}")
            return 0

    def _detect_header_row_from_df(self, df_raw: pd.DataFrame) -> int:
        """
        Detect header row from an already-loaded DataFrame (no disk re-read).

        Args:
            df_raw: DataFrame loaded with header=None

        Returns:
            0-indexed row number to use as header
        """
        try:
            if df_raw.empty:
                return 0

            best_row = 0
            best_score = 0

            for row_idx in range(min(10, len(df_raw))):
                row = df_raw.iloc[row_idx]
                score = self._score_header_row(row)

                if score > best_score:
                    best_score = score
                    best_row = row_idx

            return best_row

        except Exception:
            return 0

    def _score_header_row(self, row: pd.Series) -> int:
        """
        Score a row based on how likely it is to be a header row.

        Higher scores indicate more header-like characteristics:
        - Multiple non-empty cells
        - Contains date-like values (months, years)
        - Contains financial terms (Account, Total, etc.)
        - Has a mix of text values (not just one title cell)

        Args:
            row: A pandas Series representing a row

        Returns:
            Integer score (higher = more likely to be header)
        """
        score = 0
        non_empty_count = 0
        text_cell_count = 0
        header_pattern_matches = 0
        has_date_values = False

        for val in row:
            if pd.isna(val) or val == '' or val is None:
                continue

            non_empty_count += 1
            val_str = str(val).strip()

            # Check if it's a text value (not just a number)
            if not self._is_pure_number(val_str):
                text_cell_count += 1

            # Check for header-like patterns
            if HEADER_REGEX.search(val_str):
                header_pattern_matches += 1

            # Check for date values (pandas datetime or date-like strings)
            if hasattr(val, 'year') or self._looks_like_date(val_str):
                has_date_values = True

        # Scoring rules:
        # - Need at least 3 non-empty cells to be considered a header row
        if non_empty_count >= 3:
            score += non_empty_count * 2

        # - Bonus for having multiple text cells (headers are usually text)
        if text_cell_count >= 2:
            score += text_cell_count * 3

        # - Big bonus for matching header patterns
        score += header_pattern_matches * 10

        # - Bonus for date-like values (common in financial headers)
        if has_date_values:
            score += 5

        # - Penalty for rows that look like single title cells
        if non_empty_count == 1:
            score = 0

        return score

    def _is_pure_number(self, val_str: str) -> bool:
        """Check if a string is a pure number (int or float)."""
        try:
            float(val_str.replace(',', '').replace('$', '').replace('%', ''))
            return True
        except (ValueError, AttributeError):
            return False

    def _looks_like_date(self, val_str: str) -> bool:
        """Check if a string looks like a date."""
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',  # 2022-01-01
            r'\d{1,2}/\d{1,2}/\d{2,4}',  # 1/1/2022
            r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b',
        ]
        for pattern in date_patterns:
            if re.search(pattern, val_str, re.IGNORECASE):
                return True
        return False


def parse_excel(file_path: str) -> Dict[str, Any]:
    """
    Convenience function to parse Excel or CSV file.

    Args:
        file_path: Path to Excel/CSV file

    Returns:
        Dictionary with parsed content

    Example:
        >>> result = parse_excel("financials.xlsx")
        >>> for sheet in result['sheets']:
        ...     print(f"Sheet: {sheet['sheet_name']}, Rows: {sheet['rows']}")
    """
    parser = ExcelParser(file_path)
    return parser.parse()


def main():
    """CLI interface for testing."""
    import argparse

    parser = argparse.ArgumentParser(description="Parse Excel/CSV files")
    parser.add_argument("file", help="Path to Excel or CSV file")
    parser.add_argument("--output", help="Output JSON file path")
    parser.add_argument("--sheet", help="Specific sheet name to parse")
    parser.add_argument("--verbose", action="store_true", help="Verbose logging")

    args = parser.parse_args()

    if args.verbose:
        logger.remove()
        logger.add(sys.stderr, level="DEBUG")

    # Parse file
    result = parse_excel(args.file)

    # Output
    if args.output:
        output_path = Path(args.output)
        output_path.write_text(json.dumps(result, indent=2))
        logger.info(f"Saved results to {output_path}")
    else:
        # Print summary
        print(f"\n{'='*60}")
        print(f"File: {result['file_name']}")
        print(f"Type: {result['file_type']}")
        print(f"Sheets: {len(result['sheets'])}")
        print(f"Total Rows: {result['total_rows']:,}")
        print(f"Total Columns: {result['total_columns']}")
        print(f"Has Formulas: {result.get('has_formulas', False)}")

        if result.get('error'):
            print(f"Error: {result['error']}")

        print(f"\nSheet Details:")
        for sheet in result['sheets']:
            print(f"  - {sheet['sheet_name']}: {sheet['rows']} rows × {sheet['columns']} cols")
            if sheet.get('statistics'):
                stats = sheet['statistics']
                if stats['numeric_columns']:
                    print(f"    Numeric columns: {', '.join(stats['numeric_columns'][:5])}")

        print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
